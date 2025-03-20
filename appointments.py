import openpyxl 
import zmq 
import json
import threading
from prompt_toolkit import PromptSession
from prompt_toolkit.completion import WordCompleter
from prompt_toolkit.patch_stdout import patch_stdout


def add_appointment(new_patient):
    """Handles adding a new appointment to the excel file."""
    workbook = openpyxl.load_workbook('appointments.xlsx')
    sheet = workbook.active

    new_id = sheet.max_row
    new_patient.insert(0, new_id)

    sheet.append(new_patient)

    workbook.save('appointments.xlsx')
        
def get_appointment(search_info):
    """Handles getting the appointment for editing."""
    workbook = openpyxl.load_workbook('appointments.xlsx')
    sheet = workbook.active
    matching_rows = []

    for row_index, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
        if len(row) >= search_info[0] and row[search_info[0] - 1] == search_info[1]:
            matching_rows.append(list(row))

    return matching_rows

def get_all_appointments():
    """Handles getting all appointments."""
    try:
        workbook = openpyxl.load_workbook('appointments.xlsx')
        sheet = workbook.active  

        all_rows = []
        for row in sheet.iter_rows(values_only=True):
            all_rows.append(list(row))
        return all_rows
    except FileNotFoundError:
        print(f"Error: File not found: {'appointments.xlsx'}")
        return []
    except Exception as e:
        print(f"An error occurred: {e}")
        return []
    
def edit_appointment(appointment):
    """Handles editing a specific appointment."""
    workbook = openpyxl.load_workbook('appointments.xlsx')
    sheet = workbook.active 

    for i, data in enumerate(appointment[1]):
        sheet.cell(row=appointment[0], column=i + 1).value = data

    workbook.save(filename='appointments.xlsx')

def server_commands_interface():
    """Handles smooth shutdown of app."""
    global shutdown_flag
    global context

    commands = [
        "exit"
    ]

    completer = WordCompleter(commands)

    session = PromptSession(completer=completer)

    # Run until exit is entered or and error occurs
    while not shutdown_flag:
        with patch_stdout():
            try:
                action = session.prompt("Type exit to close 'Appointments' app: ", wrap_lines=False)
                parts = action.split()
                cmd = parts[0].lower() if parts else ""

                if cmd == "exit":
                    shutdown_flag = True
                    context.destroy()
                else:
                    print("Invalid Command!!!")

            except Exception as e:
                print(f"Error in server command interface: {e}")

if __name__ == "__main__":
    global shutdown_flag
    shutdown_flag = False

    # Create the context for communication
    global context
    context = zmq.Context()
    socket = context.socket(zmq.REP)
    socket.bind("tcp://*:5559")

    # Start the thread for the UI
    command_thread = threading.Thread(target=server_commands_interface, daemon=True)
    command_thread.start()

    # Listen for incoming messages
    try:
        while True:
            message = json.loads(socket.recv())
            print(f"Received client request: {message}")
            if len(message) > 0:
                if "add" in message:
                    add_appointment(message["add"])
                    reply = "Add Succesful"
                elif "get" in message:
                    reply = get_appointment(message["get"])
                elif "all" in message:
                    reply = get_all_appointments()
                elif "edit" in message:
                    edit_appointment(message["edit"])
                    reply = "Edit Successful"
                else:
                    reply = "Invalid Operation!!!"
                socket.send_string(json.dumps(reply))
    finally:
        print("Program terminated.\n")