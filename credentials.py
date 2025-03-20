import openpyxl 
import zmq 
import json
import threading
from prompt_toolkit import PromptSession
from prompt_toolkit.completion import WordCompleter
from prompt_toolkit.patch_stdout import patch_stdout


def add_vet(new_vet):
    """Handles adding a vet to the database."""
    workbook = openpyxl.load_workbook('vet_creds.xlsx')
    sheet = workbook.active

    new_id = sheet.max_row
    new_vet.insert(0, new_id)

    sheet.append(new_vet)

    workbook.save('vet_creds.xlsx')

def remove_vet(vet_to_remove):
    """Handles removing a vet from the database."""
    new_info = [[None, None, None]]
    new_info.insert(0, vet_to_remove + 1)
    
   
    workbook = openpyxl.load_workbook('vet_creds.xlsx')
    sheet = workbook.active  

    for i, data in enumerate(new_info[1]):
        sheet.cell(row=new_info[0], column=i + 1).value = data

    workbook.save(filename='vet_creds.xlsx')

def verify_credentials(vet_cred):
    """Handles sverification of user credentials."""
    workbook = openpyxl.load_workbook('vet_creds.xlsx')

    sheet = workbook.active

    value_to_find = vet_cred
    found = False
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == value_to_find:
                found = True
                break
        if found:
            break
    
    return found

def server_commands_interface():
    """Handles smooth shutdown of app."""
    global shutdown_flag
    global context

    commands = [
        "exit"
    ]

    completer = WordCompleter(commands)

    session = PromptSession(completer=completer)

    # Run until exit is entered or and error occurs
    while not shutdown_flag:
        with patch_stdout():
            try:
                action = session.prompt("Type exit to close 'Credentials' app: ", wrap_lines=False)
                parts = action.split()
                cmd = parts[0].lower() if parts else ""

                if cmd == "exit":
                    shutdown_flag = True
                    context.destroy()
                else:
                    print("Invalid Command!!!")

            except Exception as e:
                print(f"Error in server command interface: {e}")

if __name__ == "__main__":
    global shutdown_flag
    shutdown_flag = False

    # Create the context for communication
    global context
    context = zmq.Context()
    socket = context.socket(zmq.REP)
    socket.bind("tcp://*:5557")

    # Start the thread for the UI
    command_thread = threading.Thread(target=server_commands_interface, daemon=True)
    command_thread.start()

    # Listen for incoming messages
    try:
        while True:
            message = json.loads(socket.recv())
            print(f"Received client request: {message}")
            if len(message) > 0:
                if "add" in message:
                    add_vet(message["add"])
                    reply = "Add Succesful"
                elif "remove" in message:
                    reply = remove_vet(message["remove"])
                elif "verify" in message:
                    reply = verify_credentials(message["verify"])
                else:
                    reply = "Invalid Operation!!!"
                socket.send_string(json.dumps(reply))
    finally:
        print("Program terminated.\n")